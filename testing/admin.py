import csv
from django.http import HttpResponse
from django.contrib import admin, messages
from django.utils.html import format_html
from django.shortcuts import redirect, render
from django.urls import path
from users.models import User
from .models import (
    AppConfig, TestSuite, Question, TestResult, AIErrorLog,
    FinalQuestion, TestCategory, ReliabilityRule, LandingPage,
    UtmCampaign, UserUtm, KnowledgeBaseDocument, ProfileMatrix,
    ProfileRecommendation, Location, Payment, ReportDelivery, PromoCode
)
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.utils.safestring import mark_safe
from django.contrib.auth.forms import UserCreationForm, UserChangeForm

admin.site.site_header = "Панель управления Профтест"
admin.site.site_title = "Профтест"
admin.site.index_title = "Добро пожаловать в Профтест"
admin.site.constructor_index = admin.site.index
admin.site.constructor_app_index = admin.site.app_index


def get_test_statistics():
    from users.models import User
    from django.urls import reverse
    from django.utils import timezone
    from .models import TestResult, AppConfig, Payment
    from django.db.models import Sum

    today = timezone.now().date()
    participants = User.objects.filter(is_staff=False, is_superuser=False)

    total_users_count = participants.count()
    users_with_results_count = participants.filter(testresult__isnull=False).distinct().count()
    users_without_results_count = total_users_count - users_with_results_count
    total_tests_passed = TestResult.objects.count()

    tests_today = TestResult.objects.filter(created_at__date=today).count()
    new_users_today = participants.filter(date_joined__date=today).count()

    config = AppConfig.objects.first()
    config_url = reverse('admin:testing_appconfig_change', args=[config.pk]) if config else None

    kids = TestResult.objects.filter(child_age__lt=12).count()
    teens = TestResult.objects.filter(child_age__gte=12, child_age__lt=18).count()
    adults = TestResult.objects.filter(child_age__gte=18).count()

    total_payments = Payment.objects.filter(status='approved').count()
    total_revenue = Payment.objects.filter(status='approved').aggregate(total=Sum('amount'))['total'] or 0
    payments_today = Payment.objects.filter(status='approved', confirmed_at__date=today).count()
    revenue_today = Payment.objects.filter(status='approved', confirmed_at__date=today).aggregate(total=Sum('amount'))['total'] or 0
    pending_payments = Payment.objects.filter(status='pending').count()

    return {
        'stats': {
            'total_users': total_users_count,
            'passed_count': users_with_results_count,
            'unfinished_count': users_without_results_count,
            'total_tests_completed': total_tests_passed,
            'tests_today': tests_today,
            'new_users_today': new_users_today,
            'kids_under_12': kids,
            'teens_12_17': teens,
            'adults_18_plus': adults,
            'total_payments': total_payments,
            'total_revenue': total_revenue,
            'payments_today': payments_today,
            'revenue_today': revenue_today,
            'pending_payments': pending_payments,
        },
        'config_url': config_url
    }

class QuestionInline(admin.TabularInline):
    model = Question
    extra = 0
    fields = ('order', 'text', 'category')

class FinalQuestionInline(admin.TabularInline):
    model = FinalQuestion
    extra = 1
    fields = ('text', 'order')


@admin.register(AppConfig)
class GlobalSettingsAdmin(admin.ModelAdmin):
    def has_module_permission(self, request):
        if not request.user.is_superuser:
            return False
        return super().has_module_permission(request)

    def has_add_permission(self, request):
        exists = AppConfig.objects.exists()
        if exists:
            return False
        return True

    def has_change_permission(self, request, obj=None):
        # Только суперюзер может менять настройки
        return request.user.is_superuser

    def has_delete_permission(self, request, obj=None):
        return False

    def render_change_form(self, request, context, add=False, change=False, form_url='', obj=None):
        help_text = (
            "<b>1. Категории:</b> Создайте блоки (напр. 'Логик'). "
            "Формула: <code>([q1]+[q5])/2 * 25</code> (для 100%).<br>"
            "<b>2. Противоречивость:</b> Укажите номера через запятую (напр. <code>5,12</code>). "
            "Порог 2 — это макс. разница баллов.<br>"
            "<b>3. Однообразие:</b> Порог 0.8 — если 80% ответов одинаковые.<br>"
            "<b>4. Шкала лжи:</b> Порог 2 — если ребенок соврал в 2 и более вопросах блока.<br>"
            "<b>5. DeepSeek Лимиты:</b> Для <code>deepseek-chat</code> лимит 8K, для <code>deepseek-thinking-mode</code> лимит 64K."
        )
        context['admin_help_text'] = help_text
        return super().render_change_form(request, context, add, change, form_url, obj)

    fieldsets = (
        (None, {
            'fields': (
                'deepseek_api_key',
                'ai_max_tokens_chat',
                'ai_max_tokens_reasoner',
                'telegram_bot_token',
                'telegram_bot_name',
                'admin_notification_ids'
            )
        }),
        ('Настройки оплаты и файлы', {
            'fields': (
                'is_payment_enabled',
                'payment_price',
                'example_full_report',  # <--- ДОБАВЛЕНО ПОЛЕ
                'tochka_jwt_token',
                'tochka_customer_code',
                'tochka_merchant_id',
                'tochka_client_id',
                'tochka_client_secret'
            )
        }),
        ('Контакты поддержки', {
            'fields': (
                'support_project_name',
                'support_website',
                'support_phone',
                'support_telegram',
                'support_whatsapp',
                'support_messenger_max',
                'support_instagram',
                'support_email'
            )
        }),
    )

class TestCategoryInline(admin.TabularInline):
    model = TestCategory
    extra = 0
    fields = ('name', 'formula')

class ReliabilityRuleInline(admin.TabularInline):
    model = ReliabilityRule
    extra = 0
    fields = ('rule_type', 'question_indices', 'threshold', 'error_message', 'severity')


class ProfileMatrixInline(admin.TabularInline):
    model = ProfileMatrix
    extra = 0

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name in ["index_1", "index_2"]:
            suite_id = request.resolver_match.kwargs.get('object_id')
            if suite_id:
                kwargs["queryset"] = TestCategory.objects.filter(suite_id=suite_id)
            else:
                kwargs["queryset"] = TestCategory.objects.none()
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

class ProfileRecommendationInline(admin.StackedInline):
    model = ProfileRecommendation
    extra = 0

@admin.register(TestCategory)
class TestCategoryAdmin(admin.ModelAdmin):
    list_display = ('name', 'suite', 'formula')
    list_filter = ('suite',)
    search_fields = ('name',)


@admin.register(TestSuite)
class TestSuiteAdmin(admin.ModelAdmin):
    list_display = ('name', 'version', 'is_active', 'is_admin_only', 'ai_model_report', 'ai_temperature')
    list_filter = ('version', 'is_active', 'is_admin_only')
    inlines = [
        TestCategoryInline,
        ReliabilityRuleInline,
        QuestionInline,
        FinalQuestionInline,
        ProfileMatrixInline,
        ProfileRecommendationInline
    ]
    save_on_top = True

    fieldsets = (
        (None, {'fields': ('name', 'version', 'is_active', 'is_admin_only')}),
        ('Примеры отчетов (PDF)', {
            'fields': ('example_pdf_male', 'example_pdf_female')
        }),
        ('Настройки AI (Глобально)', {
            'fields': ('ai_model_profile', 'ai_model_report', 'ai_temperature')
        }),
        ('Системные инструкции (System)', {
            'classes': ('collapse',),
            'fields': ('profile_system_prompt', 'report_system_prompt'),
        }),
        ('Промпты пользователя (User)', {
            'fields': ('profile_prompt', 'short_result_prompt', 'extended_result_prompt', 'manager_prompt'),
        }),
        ('Контекст данных (Checklist)', {
            'description': "Выберите, что ИИ получит в системном сообщении автоматически.",
            'fields': (
                'send_child_name', 'send_child_info', 'send_location', 'send_indices', 'send_thinking_profile',
                'send_extra_data', 'send_reliability', 'send_answers_data', 'send_rag_context',
                'send_school_info', 'send_parent_info', 'send_project_info', 'send_test_meta'
            )
        }),
        ('Интерфейс', {'fields': ('description', 'closing_text', 'scale_config')}),
    )

    def questions_count(self, obj):
        return obj.questions.count()

    questions_count.short_description = "Вопросы"

    def final_questions_count(self, obj):
        return obj.final_questions.count()

    final_questions_count.short_description = "Финальные"

    class Media:
        js = ('js/question_order.js?v=1.2',)

@admin.register(Question)
class QuestionAdmin(admin.ModelAdmin):
    list_display = ('text', 'suite', 'category', 'order')
    list_filter = ('suite', 'category')
    search_fields = ('text',)
    list_editable = ('order', 'category')


class AgeGroupFilter(admin.SimpleListFilter):
    title = 'Возрастная группа'
    parameter_name = 'age_group'

    def lookups(self, request, model_admin):
        return (
            ('kids', 'Дети (до 12 лет)'),
            ('teens', 'Подростки (12-17 лет)'),
            ('adults', 'Взрослые (18+ лет)'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'kids':
            return queryset.filter(child_age__lt=12)
        if self.value() == 'teens':
            return queryset.filter(child_age__gte=12, child_age__lt=18)
        if self.value() == 'adults':
            return queryset.filter(child_age__gte=18)


@admin.register(TestResult)
class ResultsAdmin(admin.ModelAdmin):
    list_display = ('id', 'payment_status', 'child_name', 'user_phone', 'gender_short', 'child_age', 'thinking_profile_display', 'school_number', 'grade', 'reliability_status', 'report_delivery_status', 'contact_method_display', 'created_at', 'ai_actions')
    list_display_links = ('id', 'child_name')
    search_fields = ('id', 'child_name', 'school_number', 'grade', 'parent_name', 'user__phone', 'thinking_profile',
                     'contact_detail')
    search_help_text = "Поиск по ID, имени ребенка, школе, классу, родителю, профилю или контактам"
    list_filter = (
        'suite',
        'reliability_level',
        AgeGroupFilter,
        'child_gender',
        'thinking_profile',
        'contact_method',
    )
    date_hierarchy = 'created_at'

    readonly_fields = (
        'display_full_info',
        'display_profile',
        'display_manager_info',
        'display_answers',
        'display_indices',
        'display_short',
        'display_extended',
        'contact_method_display',
        'contact_detail',
        'created_at',
        'telegram_message_button'
    )
    exclude = (
        'child_name', 'child_gender', 'child_dob', 'child_age',
        'school_number', 'grade', 'parent_name', 'parent_phone',
        'answers_data', 'calculated_indices', 'extra_stats',
        'short_interpretation', 'extended_interpretation',
        'reliability_level', 'reliability_notes',
        'manager_info'
    )

    fieldsets = (
        ('Основная информация', {'fields': ('display_full_info',)}),
        ('Для менеджера', {'fields': ('display_manager_info',)}),
        ('Профиль мышления', {'fields': ('thinking_profile', 'display_profile')}),
        ('Результаты и отчеты', {
            'classes': ('tabbed',),
            'fields': ('display_indices', 'display_short', 'display_extended')
        }),
        ('Коммуникация', {'fields': ('telegram_message_button', 'contact_method_display', 'contact_detail')}),
        ('Технические данные', {
            'classes': ('collapse',),
            'fields': ('display_answers', 'created_at')
        }),
    )
    actions = ['export_to_xlsx']

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if not request.user.is_superuser:
            if self.fieldsets:
                for fieldset in self.fieldsets:
                    fields = fieldset[1].get('fields', [])
                    if isinstance(fields, tuple) or isinstance(fields, list):
                        for field in fields:
                            if field not in readonly:
                                readonly.append(field)
            if 'thinking_profile' not in readonly:
                readonly.append('thinking_profile')
        return readonly

    def has_delete_permission(self, request, obj=None):
        if not request.user.is_superuser:
            return False
        return super().has_delete_permission(request, obj)

    def gender_short(self, obj):
        return "М" if obj.child_gender == 'male' else "Ж"

    gender_short.short_description = "Пол"

    def contact_method_display(self, obj):
        has_tg = hasattr(obj.user, 'telegram_chat_id') and obj.user.telegram_chat_id
        if obj.contact_method == 'telegram' or has_tg:
            return format_html('<span style="color: #2563eb; font-weight: bold;">✅ Telegram (Бот активен)</span>')
        method_map = {'email': f'📧 Email: {obj.contact_detail}', 'max': f'💬 MAX: {obj.contact_detail}',
                      'view_only': '🛑 Только просмотр'}
        val = method_map.get(obj.contact_method, obj.contact_method)
        if obj.contact_method in ['email', 'max']:
            return format_html('<span style="color: #d97706; font-weight: bold;">{}</span>', val)
        return val

    contact_method_display.short_description = "Запрос отчета"

    def thinking_profile_display(self, obj):
        if not obj.thinking_profile: return "—"
        return format_html('<span style="color: #2563eb; font-weight: bold;">{}</span>', obj.thinking_profile)

    thinking_profile_display.short_description = "Профиль"
    thinking_profile_display.admin_order_field = 'thinking_profile'

    def user_phone(self, obj):
        return obj.user.phone

    user_phone.short_description = "📱 Номер"
    user_phone.admin_order_field = 'user__phone'

    def report_delivery_status(self, obj):
        last_full = obj.deliveries.filter(report_type='full').first()
        if last_full:
            method_label = last_full.get_method_display()
            dt = last_full.delivered_at.strftime('%d.%m %H:%M')
            return format_html(
                '<span style="color:#16a34a; font-weight:bold;" title="{}">📤 {}</span>',
                method_label, dt
            )
        return format_html('<span style="color:#dc2626; font-weight:bold;">❌</span>')

    report_delivery_status.short_description = "Отчёт"

    def reliability_status(self, obj):
        colors = {'green': '#16a34a', 'yellow': '#ca8a04', 'red': '#dc2626'}
        return format_html('<span style="color: {}; font-weight: 900;">● {}</span>',
                           colors.get(obj.reliability_level, '#9ca3af'),
                           obj.reliability_level.upper() if obj.reliability_level else "—")

    reliability_status.short_description = "Достоверность"

    def telegram_message_button(self, obj):
        from django.urls import reverse
        from django.utils.html import format_html
        if not obj.pk: return "Доступно после сохранения"
        chat_id = getattr(obj.user, 'telegram_chat_id', None)
        if not chat_id: return format_html(
            '<span style="color: #9ca3af; font-weight: bold;">❌ Telegram (Родителя) не подключен</span>')
        url = reverse('admin:user_personal_mailing', args=[obj.user.pk])
        return format_html(
            '<a href="{}" class="button" style="background-color: #2563eb; color: white; padding: 10px 20px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">✉️ НАПИСАТЬ СООБЩЕНИЕ(Telegram)</a>',
            url)

    telegram_message_button.short_description = "Отправка сообщения"
    telegram_message_button.allow_tags = True

    def payment_status(self, obj):
        if obj.is_paid:
            payment = obj.payments.filter(status='approved').first()
            if payment and payment.confirmed_at:
                from django.utils import timezone
                dt = timezone.localtime(payment.confirmed_at)
                date_str = dt.strftime('%d.%m %H:%M')
                return format_html(
                    '<div style="text-align:center;line-height:1.4;">'
                    '<span style="font-size:16px;" title="Оплачено">💰</span>'
                    '<br><span style="font-size:11px;color:#16a34a;white-space:nowrap;">{}</span>'
                    '</div>',
                    date_str
                )
            return format_html('<span style="font-size:16px;" title="Оплачено">💰</span>')
        return format_html('<span style="font-size:16px;opacity:0.3;" title="Не оплачено">⏳</span>')

    payment_status.short_description = "Оплата"
    payment_status.admin_order_field = 'is_paid'

    def display_manager_info(self, obj):
        from django.urls import reverse
        from django.utils.safestring import mark_safe
        from django.utils.html import format_html
        import markdown

        if not obj.pk:
            return "Будет доступно после сохранения."

        gen_url = reverse('admin:generate-manager-info', args=[obj.pk])
        status_url_template = reverse('admin:check-task-status', args=['TASK_ID'])

        custom_css = """
        <style>
            .ai-manager-container { 
                background:#fff1f2 !important; 
                padding:20px; 
                border-radius:12px; 
                border:2px solid #fda4af; 
                margin-bottom:15px; 
                color: #881337 !important;
                font-size: 14px;
                line-height: 1.6;
            }
            .ai-manager-container strong { font-weight: 900; color: #be123c; }
            .ai-manager-container ul { list-style-type: disc !important; margin-left: 20px !important; margin-bottom: 10px; }
            .ai-manager-container ol { list-style-type: decimal !important; margin-left: 20px !important; margin-bottom: 10px; }
            .ai-manager-container li { margin-bottom: 4px; }
            .ai-manager-container p { margin-bottom: 10px; }
            .ai-manager-container h1, .ai-manager-container h2, .ai-manager-container h3 {
                font-weight: 800; margin-top: 15px; margin-bottom: 10px; color: #9f1239;
            }
            .ai-manager-actions { margin-top: 10px; }
            .ai-manager-btn { background: #e11d48; color: white !important; padding: 8px 16px; border-radius: 6px; font-weight: bold; text-decoration: none; border: none; cursor: pointer; }
            .ai-manager-btn:hover { background: #be123c; }
        </style>
        """

        js_logic = f"""
        <script>
        function startManagerUpdate() {{
            const overlay = document.getElementById('ai-modal-overlay');
            const modalText = overlay.querySelector('.ai-modal-text');
            const modalSub = overlay.querySelector('.ai-modal-subtext');
            if(modalText) modalText.innerText = 'Генерируем инфо для менеджера';
            if(modalSub) modalSub.innerText = 'ИИ анализирует профиль и отчеты...';
            if(overlay) overlay.style.display = 'flex';

            fetch('{gen_url}', {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }})
            .then(r => r.json())
            .then(data => {{
                const taskId = data.task_id;
                const statusUrl = '{status_url_template}'.replace('TASK_ID', taskId);
                const checkStatus = () => {{
                    fetch(statusUrl, {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }})
                    .then(r => r.json())
                    .then(statusData => {{
                        if (statusData && statusData.ready) {{ window.location.reload(); }}
                        else if (statusData) {{ setTimeout(checkStatus, 2000); }}
                    }})
                    .catch(() => setTimeout(checkStatus, 3000));
                }};
                checkStatus();
            }})
            .catch(() => {{ if(overlay) overlay.style.display = 'none'; alert('Ошибка запуска задачи'); }});
        }}
        </script>
        """

        raw_content = obj.manager_info or "Информация не сгенерирована."

        html_content = markdown.markdown(raw_content)

        return format_html(
            "{}{}<div class='ai-manager-container'>{}</div><div class='ai-manager-actions'>"
            "<button type='button' onclick='startManagerUpdate()' class='ai-manager-btn'>🔄 Обновить инфо для менеджера</button>"
            "</div>",
            mark_safe(custom_css),
            mark_safe(js_logic),
            mark_safe(html_content)
        )

    display_manager_info.short_description = "Информация для менеджера (ИИ)"

    def display_full_info(self, obj):
        from django.utils.safestring import mark_safe
        if not obj.pk: return "Данные будут доступны после сохранения результата"
        gender_map = {'male': 'Мужской', 'female': 'Женский'}
        dob_display = obj.child_dob.strftime("%d.%m.%Y") if obj.child_dob else "Не указана"
        has_tg = hasattr(obj.user, 'telegram_chat_id') and obj.user.telegram_chat_id
        if obj.contact_method == 'telegram' or has_tg:
            contact_request = "Telegram (Бот подключен)"
        elif obj.contact_method in ['email', 'max']:
            contact_request = f"{obj.get_contact_method_display()}: {obj.contact_detail}"
        else:
            contact_request = "Только просмотр (Бот не подключен)"

        colors = {'green': '#16a34a', 'yellow': '#ca8a04', 'red': '#dc2626'}
        rel_color = colors.get(obj.reliability_level, '#9ca3af')
        rel_text = obj.reliability_level.upper() if obj.reliability_level else "НЕ ОПРЕДЕЛЕНА"
        reliability_html = f'<span style="color: {rel_color}; font-weight: 900;">● {rel_text}</span>'

        pay_html = "💰 <span style='color: #16a34a;'>Оплачено</span>" if obj.is_paid else "⏳ <span style='color: #6b7280;'>Ожидает оплаты</span>"

        deliveries = obj.deliveries.all()[:10]
        if deliveries:
            delivery_rows = ''
            for d in deliveries:
                method_label = d.get_method_display()
                type_label = d.get_report_type_display()
                dt = d.delivered_at.strftime('%d.%m.%Y %H:%M')
                by_name = d.delivered_by.first_name or d.delivered_by.phone if d.delivered_by else '—'
                delivery_rows += f'<tr><td style="padding:4px 8px;">{type_label}</td><td style="padding:4px 8px;">{method_label}</td><td style="padding:4px 8px;">{dt}</td><td style="padding:4px 8px;">{by_name}</td></tr>'
            delivery_html = f'<table style="width:100%; border-collapse:collapse; font-size:12px;"><tr style="background:#f3f4f6;"><th style="padding:4px 8px; text-align:left;">Тип</th><th style="padding:4px 8px; text-align:left;">Способ</th><th style="padding:4px 8px; text-align:left;">Дата</th><th style="padding:4px 8px; text-align:left;">Кто</th></tr>{delivery_rows}</table>'
        else:
            delivery_html = '<span style="color:#9ca3af;">Отчёты ещё не отправлялись</span>'

        promo_html = "—"
        if obj.promo_code:
            promo_html = f"<span style='color:#7c3aed;font-weight:900;'>{obj.promo_code}</span>"

        data = [
            ("ID Результата", f"{obj.pk}"),
            ("Статус оплаты", pay_html),
            ("Промокод", promo_html),
            ("Ребёнок", f"{obj.child_name} ({gender_map.get(obj.child_gender, obj.child_gender)})"),
            ("Дата рождения", f"{dob_display} ({obj.child_age} лет)"),
            ("Учеба", f"Школа №{obj.school_number or '—'}, {obj.grade or '—'} класс"),
            ("Родитель", f"{obj.parent_name or '—'}"),
            ("Номер ребёнка (аккаунт)", f"{obj.user.phone}"),
            ("Номер родителя", f"{obj.parent_phone or '—'}"),
            ("Профиль",
             f"<span style='color:#2563eb; font-weight:900;'>{obj.thinking_profile or 'Не определен'}</span>"),
            ("Запрос отчета", contact_request),
            ("Достоверность", reliability_html)
        ]
        html = """<div style="background: #ffffff; border: 2px solid #e5e7eb; border-radius: 16px; padding: 20px; max-width: 800px; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1);"><div style="display: grid; grid-template-columns: repeat(auto-fill, minmax(250px, 1fr)); gap: 20px;">"""
        for label, value in data:
            html += f"""<div style="border-bottom: 1px solid #f3f4f6; padding-bottom: 10px;"><div style="font-size: 10px; font-weight: 900; color: #9ca3af; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px;">{label}</div><div style="font-size: 14px; font-weight: 700; color: #111827;">{value}</div></div>"""
        html += """</div><div style="margin-top: 15px; padding-top: 15px; border-top: 2px solid #f3f4f6;"><div style="font-size: 10px; font-weight: 900; color: #9ca3af; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px;">Заметки по достоверности</div><div style="font-size: 13px; color: #374151; font-weight: 600;">{notes}</div></div>
        <div style="margin-top: 15px; padding-top: 15px; border-top: 2px solid #f3f4f6;"><div style="font-size: 10px; font-weight: 900; color: #9ca3af; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 4px;">История отправки отчётов</div><div style="margin-top:8px;">{deliveries}</div></div></div>""".format(
            notes=obj.reliability_notes or "Замечаний нет",
            deliveries=delivery_html)
        return mark_safe(html)

    display_full_info.short_description = "Полная информация"

    def display_indices(self, obj):
        if not obj.calculated_indices: return "Нет данных"
        rows = ""
        for k, v in obj.calculated_indices.items():
            suffix = "%"
            if k == "Предпочитаемый стиль работы":
                suffix = ""
            elif k == "Автономность":
                suffix = ""

            rows += f"<li style='margin-bottom:8px; color:#1a1a1a !important; font-weight:700; font-size:14px; list-style:disc; margin-left:20px;'><b>{k}</b>: <span style='color:#16a34a !important;'>{v}{suffix}</span></li>"
        return format_html(
            "<div style='background:#ffffff !important; padding:20px; border-radius:12px; border:2px solid #e5e7eb;'><ul style='margin:0; padding:0;'>{}</ul></div>",
            mark_safe(rows))

    display_indices.short_description = "Рассчитанные индексы"

    def display_answers(self, obj):
        if not obj.answers_data and not obj.extra_stats and not obj.calculated_indices: return "Нет технических данных"
        html = "<div style='background:#ffffff !important; padding:15px; border-radius:15px; border:2px solid #e5e7eb;'>"
        if obj.answers_data:
            html += "<h4 style='margin-top:0; color:#374151; font-size:12px; text-transform:uppercase;'>Основные ответы</h4><table style='width:100%; border-collapse:collapse; margin-bottom:20px;'><tr style='background:#f3f4f6;'><th style='width:30px; text-align:center; padding:8px; border:1px solid #e5e7eb; font-size:11px;'>№</th><th style='text-align:left; padding:8px; border:1px solid #e5e7eb; font-size:11px;'>Вопрос</th><th style='text-align:center; padding:8px; border:1px solid #e5e7eb; font-size:11px;'>Балл</th></tr>"
            for i, item in enumerate(obj.answers_data, 1):
                html += format_html(
                    "<tr><td style='border:1px solid #e5e7eb; padding:8px; text-align:center; font-size:11px; color:#6b7280;'>{}</td><td style='border:1px solid #e5e7eb; padding:8px; font-size:13px; font-weight:600;'>{}</td><td style='border:1px solid #e5e7eb; padding:8px; font-weight:900; text-align:center;'>{}</td></tr>",
                    i, item.get('question'), item.get('score'))
            html += "</table>"
        if obj.extra_stats:
            html += "<h4 style='color:#374151; font-size:12px; text-transform:uppercase;'>Дополнительные вопросы (Финальные)</h4><table style='width:100%; border-collapse:collapse; margin-bottom:20px;'><tr style='background:#f3f4f6;'><th style='text-align:left; padding:8px; border:1px solid #e5e7eb; font-size:11px;'>Вопрос</th><th style='text-align:left; padding:8px; border:1px solid #e5e7eb; font-size:11px;'>Ответ пользователя</th></tr>"
            for q_text, a_text in obj.extra_stats.items():
                html += format_html(
                    "<tr><td style='border:1px solid #e5e7eb; padding:8px; font-size:13px; font-weight:600;'>{}</td><td style='border:1px solid #e5e7eb; padding:8px; font-size:13px; color:#2563eb; font-weight:700;'>{}</td></tr>",
                    q_text, a_text)
            html += "</table>"
        if obj.calculated_indices:
            html += "<h4 style='color:#374151; font-size:12px; text-transform:uppercase;'>Рассчитанные индексы (Сырые данные)</h4><div style='display:flex; flex-wrap:wrap; gap:10px;'>"
            for k, v in obj.calculated_indices.items():
                if k == "Предпочитаемый стиль работы":
                    style_css, val_display = "background:#eff6ff; border:1px solid #bfdbfe; color:#1e40af;", f"{v}"
                elif k == "Автономность":
                    style_css, val_display = "background:#f3f4f6; border:1px solid #d1d5db; color:#4b5563;", f"{v} (балл)"
                else:
                    style_css, val_display = "background:#f0fdf4; border:1px solid #bbf7d0; color:#166534;", f"{v}%"
                html += format_html(
                    "<div style='padding:5px 10px; border-radius:6px; {};'><span style='font-size:11px; font-weight:700; opacity:0.8;'>{}:</span> <span style='font-size:12px; font-weight:900;'>{}</span></div>",
                    style_css, k, val_display)
            html += "</div>"
        html += "</div>"
        return mark_safe(html)

    display_answers.short_description = "Данные ответов"

    def display_profile(self, obj):
        from django.urls import reverse
        from django.utils.safestring import mark_safe
        from django.utils.html import format_html
        if not obj.pk: return "Будет доступно после сохранения."
        gen_url = reverse('admin:generate-ai-profile-only', args=[obj.pk])
        download_url = reverse('admin:download-pdf-report', args=[obj.pk]) + "?type=profile"
        status_url_template = reverse('admin:check-task-status', args=['TASK_ID'])
        custom_css = """<style>.ai-profile-container { background:#f8fafc !important; padding:20px; border-radius:12px; border:2px solid #e2e8f0; margin-bottom:15px; }.ai-profile-value { font-size: 18px; font-weight: 800; color: #1e40af !important; margin-bottom: 10px; display: block; }.ai-profile-actions { display: flex; gap: 10px; flex-wrap: wrap; }</style>"""
        js_logic = f"""<script>function startProfileUpdate() {{ const overlay = document.getElementById('ai-modal-overlay'); const modalText = overlay.querySelector('.ai-modal-text'); const modalSub = overlay.querySelector('.ai-modal-subtext'); modalText.innerText = 'Определяем профиль'; modalSub.innerText = 'ИИ анализирует индексы и ответы...'; overlay.style.display = 'flex'; fetch('{gen_url}', {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(data => {{ const taskId = data.task_id; const statusUrl = '{status_url_template}'.replace('TASK_ID', taskId); const checkStatus = () => {{ fetch(statusUrl, {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(statusData => {{ if (statusData && statusData.ready) {{ window.location.reload(); }} else if (statusData) {{ setTimeout(checkStatus, 2000); }} }}).catch(() => setTimeout(checkStatus, 3000)); }}; checkStatus(); }}).catch(() => {{ overlay.style.display = 'none'; alert('Ошибка запуска задачи'); }}); }}</script>"""
        profile_html = f"<span class='ai-profile-value'>{obj.thinking_profile or 'Не определен'}</span>"
        return format_html(
            "{}{}<div class='ai-profile-container'>{}<div class='ai-profile-actions'><button type='button' onclick='startProfileUpdate()' class='ai-update-btn' style='background:#6366f1;'>Обновить профиль</button><a href='{}' target='_blank' class='ai-download-btn' style='background:#64748b;'>Скачать PDF (Профиль)</a></div></div>",
            mark_safe(custom_css), mark_safe(js_logic), mark_safe(profile_html), download_url)

    display_profile.short_description = "Профиль мышления (ИИ)"

    def display_short(self, obj):
        from django.urls import reverse
        from django.utils.safestring import mark_safe
        from django.utils.html import format_html
        if not obj.pk: return "Будет доступно после сохранения."
        gen_url = reverse('admin:generate-short-ai-report', args=[obj.pk])
        send_tg_url = reverse('admin:send-short-report-tg', args=[obj.pk])
        download_url = reverse('admin:download-pdf-report', args=[obj.pk]) + "?type=short"
        status_url_template = reverse('admin:check-task-status', args=['TASK_ID'])
        show_tg_btn = bool(obj.user.telegram_chat_id)
        tg_btn_html = ""
        if show_tg_btn: tg_btn_html = f"<button type='button' id='short-tg-btn' onclick='sendShortPdf()' class='ai-sms-btn'>Отправить PDF (Предв.)</button>"
        custom_css = "<style>.ai-short-container { background:#ffffff !important; color:#1a1a1a !important; padding:30px; border-radius:20px; border:3px solid #d1d5db; line-height:1.8; font-size:16px; font-weight:600; margin-bottom: 20px; } .ai-short-container * { color:#1a1a1a !important; } .ai-short-container p { margin-bottom:15px !important; } .ai-btn-group { display: flex; gap: 15px; margin-top: 10px; justify-content: flex-start; flex-wrap: wrap; } .ai-update-btn, .ai-sms-btn, .ai-download-btn { display: inline-flex; align-items: center; justify-content: center; padding: 10px 20px; border-radius: 6px; font-weight: 700; font-size: 13px; border: none; cursor: pointer; transition: all 0.2s; text-transform: uppercase; color: white !important; text-decoration: none; } .ai-update-btn { background: #16a34a; } .ai-update-btn:hover { background: #15803d; } .ai-sms-btn { background: #3b82f6; } .ai-sms-btn:hover { background: #2563eb; } .ai-download-btn { background: #6b7280; } .ai-download-btn:hover { background: #4b5563; }</style>"
        js_logic = f"""<script>function startShortUpdate() {{ const overlay = document.getElementById('ai-modal-overlay'); const modalText = overlay.querySelector('.ai-modal-text'); const modalSub = overlay.querySelector('.ai-modal-subtext'); modalText.innerText = 'Переписываем краткий отчет'; modalSub.innerText = 'ИИ анализирует данные заново...'; overlay.style.display = 'flex'; fetch('{gen_url}', {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(data => {{ const taskId = data.task_id; const statusUrl = '{status_url_template}'.replace('TASK_ID', taskId); const checkStatus = () => {{ fetch(statusUrl, {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(statusData => {{ if (statusData && statusData.ready) {{ window.location.reload(); }} else if (statusData) {{ setTimeout(checkStatus, 2000); }} }}).catch(() => setTimeout(checkStatus, 3000)); }}; checkStatus(); }}).catch(() => {{ overlay.style.display = 'none'; alert('Ошибка запуска задачи'); }}); }} function sendShortPdf() {{ const btn = document.getElementById('short-tg-btn'); const originalText = btn.innerText; btn.innerText = 'ОТПРАВЛЯЮ...'; btn.disabled = true; fetch('{send_tg_url}', {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(data => {{ if (data.status === 'ok') {{ alert('Предварительный PDF отправлен!'); }} else {{ alert('Ошибка: ' + data.message); }} }}).catch(() => alert('Ошибка при отправке')).finally(() => {{ btn.innerText = originalText; btn.disabled = false; }}); }}</script>"""
        content = obj.short_interpretation or "Ожидание генерации..."
        return format_html(
            "{}{}<div class='ai-short-container'>{}</div><div class='ai-btn-group'><button type='button' onclick='startShortUpdate()' class='ai-update-btn'>Обновить Краткий Отчет</button>{}<a href='{}' target='_blank' class='ai-download-btn'>Скачать PDF (Краткий)</a></div>",
            mark_safe(custom_css), mark_safe(js_logic), mark_safe(content), mark_safe(tg_btn_html), download_url)

    display_short.short_description = "Краткое заключение (ИИ)"

    def display_extended(self, obj):
        from django.urls import reverse
        from django.utils.safestring import mark_safe
        from django.utils.html import format_html
        if not obj.pk: return "Полный отчет будет доступен после сохранения результата и обработки данных ИИ."
        gen_url = reverse('admin:generate-ai-report', args=[obj.pk])
        sms_url = reverse('admin:send-report-sms', args=[obj.pk])
        download_url = reverse('admin:download-pdf-report', args=[obj.pk])
        status_url_template = reverse('admin:check-task-status', args=['TASK_ID'])
        show_tg_btn = True
        if not obj.user.telegram_chat_id: show_tg_btn = False
        tg_btn_html = ""
        if show_tg_btn: tg_btn_html = f"<button type='button' id='sms-report-btn' onclick='sendTelegramPdf()' class='ai-sms-btn'>Отправить PDF в Telegram</button>"
        contact_info_html = ""
        if obj.contact_method == 'email':
            contact_info_html = f"<div style='margin-bottom:20px; padding:10px; background:#fff7ed; border-left:4px solid #f97316; color:#9a3412;'><b>Запрос на Email:</b> {obj.contact_detail}. <br>Сформируйте PDF и отправьте вручную.</div>"
        elif obj.contact_method == 'max':
            contact_info_html = f"<div style='margin-bottom:20px; padding:10px; background:#fff7ed; border-left:4px solid #f97316; color:#9a3412;'><b>Запрос в MAX:</b> {obj.contact_detail}. <br>Сформируйте PDF и отправьте вручную.</div>"
        custom_style = """<style>.ai-extended-wrapper { background: #ffffff !important; padding: 30px; border-radius: 16px; border: 1px solid #e5e7eb; box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1); line-height: 1.6; font-size: 16px; color: #1f2937 !important; width: 100%; box-sizing: border-box; margin-top: 15px; font-family: ui-sans-serif, system-ui, -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif; } .ai-extended-wrapper * { color: #1f2937 !important; box-sizing: border-box; } .ai-extended-wrapper h1 { margin-top: 0 !important; margin-bottom: 20px !important; font-size: 24px !important; font-weight: 800 !important; color: #111827 !important; border-bottom: 2px solid #3b82f6; padding-bottom: 10px; } .ai-extended-wrapper h2 { margin-top: 30px !important; margin-bottom: 15px !important; font-size: 20px !important; font-weight: 700 !important; color: #1e40af !important; background-color: #f0f9ff !important; padding: 10px; border-radius: 6px; } .ai-extended-wrapper h3 { margin-top: 20px !important; margin-bottom: 10px !important; font-size: 18px !important; font-weight: 600 !important; color: #374151 !important; } .ai-extended-wrapper p { margin-bottom: 12px !important; text-align: left; } .ai-extended-wrapper ul, .ai-extended-wrapper ol { margin-left: 20px !important; margin-bottom: 15px !important; } .ai-extended-wrapper li { margin-bottom: 6px !important; padding-left: 5px; } .ai-extended-wrapper table { width: 100% !important; border-collapse: collapse; margin-bottom: 20px; background-color: #ffffff !important; } .ai-extended-wrapper tr { background-color: #ffffff !important; } .ai-extended-wrapper td, .ai-extended-wrapper th { border: 1px solid #e5e7eb; padding: 10px; text-align: left; background-color: #ffffff !important; color: #1f2937 !important; } .ai-btn-group { display: flex; gap: 15px; margin-top: 30px; justify-content: flex-start; flex-wrap: wrap; } .ai-update-btn, .ai-sms-btn, .ai-download-btn { display: inline-flex; align-items: center; justify-content: center; padding: 10px 20px; text-decoration: none; border-radius: 6px; font-weight: 700; font-size: 13px; border: none; cursor: pointer; transition: all 0.2s; text-transform: uppercase; } .ai-update-btn { background: #16a34a; color: white !important; } .ai-update-btn:hover { background: #15803d; } .ai-sms-btn { background: #3b82f6; color: white !important; } .ai-sms-btn:hover { background: #2563eb; } .ai-download-btn { background: #6b7280; color: white !important; } .ai-download-btn:hover { background: #4b5563; } #ai-modal-overlay { position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(15, 23, 42, 0.8); display: none; align-items: center; justify-content: center; z-index: 10000; backdrop-filter: blur(4px); } .ai-modal-content { background: white; padding: 40px; border-radius: 24px; text-align: center; max-width: 400px; width: 90%; box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.1); } .ai-spinner { width: 50px; height: 50px; border: 4px solid #f1f5f9; border-top: 4px solid #3b82f6; border-radius: 50%; animation: ai-spin 1s linear infinite; margin: 0 auto 20px; } @keyframes ai-spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } } .ai-modal-text { color: #111827 !important; font-weight: 700; font-size: 20px; margin-bottom: 8px; } .ai-modal-subtext { color: #6b7280 !important; font-size: 14px; }</style>"""
        js_logic = f"""<script>function startAiUpdate() {{ const overlay = document.getElementById('ai-modal-overlay'); const modalText = overlay.querySelector('.ai-modal-text'); const modalSub = overlay.querySelector('.ai-modal-subtext'); modalText.innerText = 'Переписываем отчет'; modalSub.innerText = 'ИИ анализирует данные заново. Модальное окно закроется автоматически.'; overlay.style.display = 'flex'; fetch('{gen_url}', {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(data => {{ const taskId = data.task_id; const statusUrl = '{status_url_template}'.replace('TASK_ID', taskId); const checkStatus = () => {{ fetch(statusUrl, {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(statusData => {{ if (statusData && statusData.ready) {{ window.location.reload(); }} else if (statusData) {{ setTimeout(checkStatus, 2000); }} }}).catch(() => setTimeout(checkStatus, 3000)); }}; checkStatus(); }}).catch(() => {{ overlay.style.display = 'none'; alert('Ошибка запуска задачи'); }}); }} function sendTelegramPdf() {{ const btn = document.getElementById('sms-report-btn'); const originalText = btn.innerText; btn.innerText = 'ОТПРАВЛЯЮ...'; btn.disabled = true; fetch('{sms_url}', {{ headers: {{ 'X-Requested-With': 'XMLHttpRequest' }} }}).then(r => r.json()).then(data => {{ if (data.status === 'ok') {{ alert('PDF успешно отправлен в Telegram!'); }} else {{ alert('Ошибка: ' + data.message); }} }}).catch(() => alert('Ошибка при отправке')).finally(() => {{ btn.innerText = originalText; btn.disabled = false; }}); }}</script>"""
        modal_html = """<div id="ai-modal-overlay"><div class="ai-modal-content"><div class="ai-spinner"></div><div class="ai-modal-text"></div><div class="ai-modal-subtext"></div></div></div>"""
        content = obj.extended_interpretation or "Отчет еще не сформирован."
        return format_html(
            "{}{}{}{}<div class='ai-extended-wrapper'>{}<div class='ai-btn-group'><button type='button' onclick='startAiUpdate()' class='ai-update-btn'>Обновить отчет</button>{}<a href='{}' target='_blank' class='ai-download-btn'>Скачать PDF</a></div></div>",
            mark_safe(custom_style), mark_safe(js_logic), mark_safe(modal_html), mark_safe(contact_info_html),
            mark_safe(content), mark_safe(tg_btn_html), download_url)

    display_extended.short_description = "Полный отчет (ИИ)"

    def ai_actions(self, obj):
        from django.urls import reverse
        if not obj.pk: return "—"
        url = reverse('admin:generate-ai-report', args=[obj.pk])
        return format_html(
            '<a class="button" style="background: #16a34a; color: white; padding: 6px 12px; text-decoration: none; border-radius: 6px; font-weight: 900; text-transform: uppercase; font-size: 10px; white-space: nowrap; display: inline-block; line-height: 1.2;" href="{}">Обновить ИИ</a>',
            url)

    ai_actions.short_description = "Действия ИИ"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('<int:pk>/ai-gen/', self.admin_site.admin_view(self.generate_ai_report), name='generate-ai-report'),
            path('<int:pk>/ai-gen-short/', self.admin_site.admin_view(self.generate_short_ai_report),
                 name='generate-short-ai-report'),
            path('<int:pk>/ai-gen-profile/', self.admin_site.admin_view(self.generate_ai_profile_view),
                 name='generate-ai-profile-only'),
            path('<int:pk>/ai-gen-manager/', self.admin_site.admin_view(self.generate_manager_info_view),
                 name='generate-manager-info'),
            path('check-task/<str:task_id>/', self.admin_site.admin_view(self.check_task_status),
                 name='check-task-status'),
            path('<int:pk>/send-sms/', self.admin_site.admin_view(self.send_report_sms), name='send-report-sms'),
            path('<int:pk>/send-short-tg/', self.admin_site.admin_view(self.send_short_report_tg_view),
                 name='send-short-report-tg'),
            path('<int:pk>/download-pdf/', self.admin_site.admin_view(self.download_pdf_report),
                 name='download-pdf-report'),
        ]
        return custom_urls + urls

    def generate_ai_report(self, request, pk):
        from .tasks import generate_extended_report_task
        from django.http import JsonResponse
        task = generate_extended_report_task.delay(pk)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'task_id': task.id})
        messages.info(request, "Задача отправлена.")
        return redirect('..')

    def generate_ai_profile_view(self, request, pk):
        from .tasks import generate_ai_profile_task
        from django.http import JsonResponse
        task = generate_ai_profile_task.delay(pk, is_initial=False)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'task_id': task.id})
        messages.info(request, "Задача обновления профиля запущена.")
        return redirect('..')

    def generate_short_ai_report(self, request, pk):
        from .tasks import generate_short_report_task
        from django.http import JsonResponse
        task = generate_short_report_task.delay(pk, automatic_send=False)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'task_id': task.id})
        messages.info(request, "Задача обновления краткого отчета запущена.")
        return redirect('..')

    def generate_manager_info_view(self, request, pk):
        from .tasks import generate_manager_info_task
        from django.http import JsonResponse
        task = generate_manager_info_task.delay(pk)
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'task_id': task.id})
        messages.info(request, "Задача обновления инфо для менеджера запущена.")
        return redirect('..')

    class Media:
        js = ('js/admin_top_scroll.js',)
        css = {
            'all': ('css/admin_extra_scroll.css',)
        }

    def check_task_status(self, request, task_id):
        from celery.result import AsyncResult
        from core.celery import app
        from django.http import JsonResponse
        res = AsyncResult(task_id, app=app)
        return JsonResponse({'status': res.status, 'ready': res.ready()})

    def send_report_sms(self, request, pk):
        from django.http import JsonResponse
        from .tasks import send_telegram_result_task
        from django.shortcuts import get_object_or_404
        obj = get_object_or_404(TestResult, pk=pk)
        if not obj.user.telegram_chat_id:
            return JsonResponse({'status': 'error', 'message': 'У пользователя не привязан Telegram (нет chat_id)'})
        task_res = send_telegram_result_task.delay(pk, report_type='full', delivery_method='manual_telegram', delivered_by_id=request.user.pk)
        return JsonResponse({'status': 'ok', 'task_id': str(task_res)})

    def send_short_report_tg_view(self, request, pk):
        from django.http import JsonResponse
        from django.shortcuts import get_object_or_404
        from .models import AppConfig, TestResult
        from .tasks import send_telegram_result_task
        obj = get_object_or_404(TestResult, pk=pk)
        if not obj.user.telegram_chat_id:
            return JsonResponse({'status': 'error', 'message': 'У пользователя нет Telegram ID'})
        config = AppConfig.load()
        if not config.telegram_bot_token:
            return JsonResponse({'status': 'error', 'message': 'Бот не настроен'})
        task_res = send_telegram_result_task.delay(pk, report_type='short', delivery_method='manual_telegram', delivered_by_id=request.user.pk)
        return JsonResponse({'status': 'ok', 'task_id': str(task_res)})

    def download_pdf_report(self, request, pk):
        from django.http import FileResponse
        from django.shortcuts import get_object_or_404
        from .services import PDFService
        from .models import TestResult
        obj = get_object_or_404(TestResult, pk=pk)
        report_type = request.GET.get('type', 'full')
        pdf_buffer = PDFService.generate_pdf(obj, report_type=report_type)
        safe_name = str(obj.child_name).replace(" ", "_")
        prefix = "Short" if report_type == 'short' else "Full"
        filename = f"{prefix}_Result_{safe_name}.pdf"
        return FileResponse(pdf_buffer, as_attachment=True, filename=filename)

    def export_to_xlsx(self, request, queryset):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
        import re
        from django.utils import timezone
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты тестирования"

        # ── Стили ──
        thin_border = Border(
            left=Side(style='thin', color='D1D5DB'),
            right=Side(style='thin', color='D1D5DB'),
            top=Side(style='thin', color='D1D5DB'),
            bottom=Side(style='medium', color='9CA3AF')
        )
        wrap_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
        center_align = Alignment(horizontal="center", vertical="top", wrap_text=False)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

        # Цвета блоков заголовков
        BLOCK_COLORS = {
            'main': PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid"),
            'contacts': PatternFill(start_color="059669", end_color="059669", fill_type="solid"),
            'results': PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid"),
            'payment': PatternFill(start_color="D97706", end_color="D97706", fill_type="solid"),
            'ai': PatternFill(start_color="4B5563", end_color="4B5563", fill_type="solid"),
        }
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        data_font = Font(size=10, name="Calibri", color="000000")
        data_font_bold = Font(size=10, name="Calibri", bold=True, color="000000")

        # Зебра
        zebra_even = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        zebra_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Достоверность
        reliability_fills = {
            'green': PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
            'yellow': PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
            'red': PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        }
        paid_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        unpaid_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

        def clean_html(raw_html):
            if not raw_html:
                return ""
            clean = raw_html.replace('<br />', '\n').replace('<br>', '\n')
            clean = clean.replace('</p>', '\n\n').replace('</li>', '\n')
            clean = clean.replace('<li>', '• ')
            clean = re.sub(r'<.*?>', '', clean)
            lines = [line.strip() for line in clean.split('\n')]
            return '\n'.join(line for line in lines if line).strip()

        def age_group(age):
            if age is None:
                return "—"
            if age <= 6:
                return "до 7 лет"
            elif age <= 11:
                return "7–11 лет"
            elif age <= 14:
                return "12–14 лет"
            else:
                return "15–17 лет"


        # ── Конфигурация колонок ──
        columns = [
            # Блок 1: Основное
            ('main', '№', 10),
            ('main', 'ID', 10),
            ('main', 'Дата теста', 18),
            ('main', 'Ребёнок', 24),
            ('main', 'Пол', 10),
            ('main', 'ДР', 14),
            ('main', 'Возраст', 12),
            ('main', 'Группа', 14),
            ('main', 'Школа', 14),
            ('main', 'Класс', 10),
            ('main', 'Локация', 18),
            # Блок 2: Контакты
            ('contacts', 'Тел. ребёнка', 18),
            ('contacts', 'Родитель', 22),
            ('contacts', 'Тел. родителя', 18),
            ('contacts', 'TG', 8),
            ('contacts', 'Запрос', 14),
            ('contacts', 'Контакт', 20),
            # Блок 3: Результаты
            ('results', 'Профиль', 18),
            ('results', 'Достов.', 10),
            ('results', 'Заметки', 25),
            ('results', 'Индексы', 30),
            ('results', 'Интересы', 20),
            ('results', 'Хобби', 20),
            # Блок 4: Оплата и отправка
            ('payment', 'Оплата', 9),
            ('payment', 'Дата опл.', 16),
            ('payment', 'Отчёт', 9),
            ('payment', 'Дата отпр.', 16),
            ('payment', 'Способ', 14),
            # Блок 5: AI-отчёты
            ('ai', 'Менеджеру', 35),
            ('ai', 'Кратко', 40),
            ('ai', 'Полный', 50),
        ]

        # ── Шапка ──
        for col_idx, (block, title, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=title)
            cell.font = header_font
            cell.fill = BLOCK_COLORS[block]
            cell.border = thin_border
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Фиксируем шапку и высоту строки
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

        # ── Данные ──


        for row_offset, obj in enumerate(queryset):
            row_num = row_offset + 2
            row_fill = zebra_even if row_offset % 2 == 0 else zebra_odd

            # Индексы
            indices_parts = []
            if obj.calculated_indices:
                for k, v in obj.calculated_indices.items():
                    suffix = "%"
                    if k in ("Предпочитаемый стиль работы", "Автономность"):
                        suffix = ""
                    indices_parts.append(f"{k}: {v}{suffix}")
            indices_text = "\n".join(indices_parts)

            # Telegram
            tg_status = "✅" if hasattr(obj.user, 'telegram_chat_id') and obj.user.telegram_chat_id else "❌"

            # Оплата
            paid_text = "💰 Да" if obj.is_paid else "⏳ Нет"

            # Дата оплаты (из Payment модели)
            payment_date = ""
            try:
                last_payment = obj.payments.filter(status='approved').first()
                if last_payment and last_payment.confirmed_at:
                    payment_date = timezone.localtime(last_payment.confirmed_at).strftime("%d.%m.%Y %H:%M")
            except Exception:
                pass

            # Отправка отчёта (из ReportDelivery)
            report_sent = "❌ Нет"
            report_sent_date = ""
            report_sent_method = ""
            try:
                last_delivery = obj.deliveries.filter(report_type='full').first()
                if last_delivery:
                    report_sent = "📤 Да"
                    report_sent_date = timezone.localtime(last_delivery.delivered_at).strftime("%d.%m.%Y %H:%M")
                    report_sent_method = last_delivery.get_method_display()
            except Exception:
                pass

            # Достоверность
            rel_map = {'green': '🟢 Норма', 'yellow': '🟡 Внимание', 'red': '🔴 Риск'}
            reliability_text = rel_map.get(obj.reliability_level, obj.reliability_level)

            # Локация
            location_name = obj.location.name if obj.location else "—"

            # Контакт-метод
            contact_map = {'telegram': 'Telegram', 'email': 'Email', 'max': 'MAX', 'view_only': 'Только просмотр'}
            contact_method_text = contact_map.get(obj.contact_method, obj.contact_method or '')

            row_data = [
                row_offset + 1,                                                              # №
                obj.pk,                                                                       # ID
                timezone.localtime(obj.created_at).strftime("%d.%m.%Y %H:%M"),               # Дата теста
                obj.child_name,                                                               # Ребёнок
                "М" if obj.child_gender == "male" else ("Ж" if obj.child_gender == "female" else obj.child_gender),
                obj.child_dob.strftime("%d.%m.%Y") if obj.child_dob else "—",                # ДР
                obj.child_age,                                                                # Возраст
                age_group(obj.child_age),                                                     # Группа
                obj.school_number or "—",                                                     # Школа
                obj.grade or "—",                                                             # Класс
                location_name,                                                                # Локация
                # Контакты
                obj.user.phone,                                                               # Номер ребёнка
                obj.parent_name or "—",                                                       # Родитель
                obj.parent_phone or "—",                                                      # Номер родителя
                tg_status,                                                                    # Telegram
                contact_method_text,                                                          # Запрос отчёта
                obj.contact_detail or "—",                                                    # Контакт для отчёта
                # Результаты
                obj.thinking_profile or "—",                                                  # Профиль мышления
                reliability_text,                                                             # Достоверность
                obj.reliability_notes or "—",                                                 # Заметки
                indices_text or "—",                                                          # Индексы
                obj.extra_stats.get('interest_sphere', '—') if obj.extra_stats else '—',      # Интересы
                obj.extra_stats.get('hobby_text', '—') if obj.extra_stats else '—',           # Хобби
                # Оплата
                paid_text,                                                                    # Оплачен
                payment_date or "—",                                                          # Дата оплаты
                report_sent,                                                                  # Отчёт отправлен
                report_sent_date or "—",                                                      # Дата отправки
                report_sent_method or "—",                                                    # Способ отправки
                # AI
                obj.manager_info or "—",                                                      # Инфо для менеджера
                clean_html(obj.short_interpretation) or "—",                                  # Краткое заключение
                clean_html(obj.extended_interpretation) or "—",                               # Полный отчёт
            ]


            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.fill = row_fill

                col_title = columns[col_idx - 1][1] if col_idx <= len(columns) else ''

                # Выравнивание
                if col_title in ('№', 'ID', 'Возраст', 'Пол', 'TG', 'Оплата', 'Отчёт'):
                    cell.alignment = center_align
                else:
                    cell.alignment = wrap_align

                # Спец. заливки
                if col_title == 'Достов.':
                    fill = reliability_fills.get(obj.reliability_level)
                    if fill:
                        cell.fill = fill

                if col_title == 'Оплата':
                    cell.fill = paid_fill if obj.is_paid else unpaid_fill
                    cell.font = data_font_bold

                if col_title == 'Ребёнок':
                    cell.font = data_font_bold

        # ── Автофильтр ──
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

        # ── Файл ──
        now_str = timezone.localtime(timezone.now()).strftime("%d.%m.%Y_%H-%M")
        filename = f"results_{now_str}.xlsx"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    export_to_xlsx.short_description = "📊 Выгрузить в Excel (XLSX)"


class TelegramFilter(admin.SimpleListFilter):
    title = 'Telegram'
    parameter_name = 'telegram_status'

    def lookups(self, request, model_admin):
        return (
            ('connected', '✅ Подключен'),
            ('not_connected', '❌ Не подключен'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'connected':
            return queryset.exclude(telegram_chat_id__isnull=True).exclude(telegram_chat_id='')
        if self.value() == 'not_connected':
            return queryset.filter(telegram_chat_id__in=[None, ''])


class CustomUserCreationForm(UserCreationForm):
    class Meta:
        model = User
        fields = ('phone', 'first_name', 'is_staff', 'is_active', 'groups', 'consent_given')

    def save(self, commit=True):
        user = super().save(commit=False)
        if commit:
            user.save()
            self.save_m2m()
        return user


class CustomUserChangeForm(UserChangeForm):
    class Meta:
        model = User
        fields = '__all__'


@admin.register(User)
class UserAdmin(BaseUserAdmin):
    add_form = CustomUserCreationForm
    form = CustomUserChangeForm

    list_display = ('phone', 'first_name', 'get_payment_status', 'get_child_name', 'get_child_age', 'get_parent_name', 'get_test_count', 'get_utm', 'get_telegram_status', 'date_joined', 'is_staff', 'is_active')
    list_filter = ('is_staff', 'is_active', 'consent_given', 'utm_info__utm_source', TelegramFilter)

    readonly_fields = ('last_login', 'date_joined', 'telegram_message_button', 'reset_telegram_button_display')

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('phone', 'first_name', 'password1', 'password2', 'is_staff', 'is_active', 'groups',
                       'consent_given'),
        }),
    )

    fieldsets = (
        (None, {'fields': ('phone', 'password')}),
        ('Персональная информация', {'fields': ('first_name',)}),
        ('Юридическая информация', {'fields': ('consent_given', 'consent_date')}),
        ('Права доступа', {'fields': ('is_active', 'is_staff', 'is_superuser', 'groups', 'user_permissions')}),
        ('Важные даты', {'fields': ('last_login', 'date_joined')}),
        ('Действия', {'fields': ('telegram_message_button', 'reset_telegram_button_display')}),
    )

    search_fields = ('phone', 'first_name', 'testresult__child_name', 'testresult__parent_name', 'utm_info__utm_source',
                     'utm_info__utm_campaign')
    ordering = ('phone',)
    filter_horizontal = ('groups', 'user_permissions')
    actions = ['send_telegram_action', 'reset_telegram_action']

    def get_readonly_fields(self, request, obj=None):
        readonly = list(super().get_readonly_fields(request, obj))
        if obj is not None and not request.user.is_superuser:
            if self.fieldsets:
                for fieldset in self.fieldsets:
                    fields = fieldset[1].get('fields', [])
                    if isinstance(fields, (tuple, list)):
                        for field in fields:
                            if field not in readonly:
                                readonly.append(field)
            for extra in ['phone', 'password', 'is_staff', 'is_superuser', 'groups', 'user_permissions']:
                if extra not in readonly:
                    readonly.append(extra)
        return readonly

    def has_add_permission(self, request):
        if not request.user.is_superuser:
            return False
        return super().has_add_permission(request)

    def has_delete_permission(self, request, obj=None):
        if not request.user.is_superuser:
            return False
        return super().has_delete_permission(request, obj)

    def reset_telegram_action(self, request, queryset):
        if not request.user.is_superuser:
            self.message_user(request, "У вас недостаточно прав для этого действия.", level='error')
            return
        rows_updated = queryset.update(telegram_chat_id=None)
        self.message_user(request, f"Telegram успешно отвязан у {rows_updated} пользователей.")

    reset_telegram_action.short_description = "❌ Сбросить привязку Telegram"

    def get_test_count(self, obj):
        from .models import TestResult
        count = TestResult.objects.filter(user=obj).count()
        return format_html('<span style="font-weight: bold;">{}</span>', count)

    get_test_count.short_description = "Тестов"

    def get_payment_status(self, obj):
        if obj.testresult_set.filter(is_paid=True).exists():
            return format_html('<span style="font-size: 16px;" title="Есть оплаченные тесты">💰</span>')
        return format_html('<span style="font-size: 16px; opacity: 0.2;" title="Нет оплат">➖</span>')

    get_payment_status.short_description = "Оплата"

    def get_utm(self, obj):
        if hasattr(obj, 'utm_info'):
            return f"{obj.utm_info.utm_source} / {obj.utm_info.utm_campaign or '-'}"
        return "—"

    get_utm.short_description = "Источник (UTM)"

    def get_telegram_status(self, obj):
        chat_id = getattr(obj, 'telegram_chat_id', None)
        if chat_id:
            return format_html('<span style="color: #16a34a; font-weight: bold;">✅ Подключен</span>')
        return format_html('<span style="color: #cbd5e1;">—</span>')

    get_telegram_status.short_description = "Telegram"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('mailing/', self.admin_site.admin_view(self.mailing_view), name='user_mailing'),
            path('<int:object_id>/personal-mailing/', self.admin_site.admin_view(self.personal_mailing_view),
                 name='user_personal_mailing'),
            path('<int:pk>/reset-telegram/', self.admin_site.admin_view(self.reset_telegram_view),
                 name='user_reset_telegram'),
        ]
        return custom_urls + urls

    def reset_telegram_view(self, request, pk):
        from django.shortcuts import get_object_or_404, redirect
        if not request.user.is_superuser:
            self.message_user(request, "У вас нет прав для отвязки Telegram.", level='error')
            return redirect('admin:users_user_change', pk)
        user = get_object_or_404(User, pk=pk)
        user.telegram_chat_id = None
        user.save()
        self.message_user(request, f"Telegram успешно отвязан у пользователя {user.phone}.")
        return redirect('admin:users_user_change', pk)

    def reset_telegram_button_display(self, obj):
        from django.urls import reverse
        if not obj.pk or not obj.telegram_chat_id:
            return format_html('<span style="color: #9ca3af;">— Telegram не подключен —</span>')
        url = reverse('admin:user_reset_telegram', args=[obj.pk])
        return format_html(
            '<a href="{}" class="button" style="background-color: #dc2626; color: white; padding: 10px 20px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">'
            '❌ ОТВЯЗАТЬ TELEGRAM</a>',
            url
        )

    reset_telegram_button_display.short_description = "Сброс привязки"
    reset_telegram_button_display.allow_tags = True

    def send_telegram_action(self, request, queryset):
        from django.urls import reverse
        from django.shortcuts import render
        selected_ids = list(queryset.values_list('pk', flat=True))
        ajax_url = reverse('admin:user_mailing')
        context = {
            **self.admin_site.each_context(request),
            'users': queryset,
            'selected_ids': ",".join(map(str, selected_ids)),
            'opts': self.model._meta,
            'is_action': True,
            'title': 'Отправка сообщения выбранным пользователям',
            'ajax_url': ajax_url
        }
        return render(request, 'admin/mailing.html', context)

    send_telegram_action.short_description = "Отправить сообщение в Telegram"

    def get_child_name(self, obj):
        result = obj.testresult_set.last()
        return result.child_name if result else "—"

    get_child_name.short_description = "Имя ребенка"

    def get_child_age(self, obj):
        result = obj.testresult_set.last()
        return result.child_age if result else "—"

    get_child_age.short_description = "Возраст"

    def get_parent_name(self, obj):
        result = obj.testresult_set.last()
        return result.parent_name if result else "—"

    get_parent_name.short_description = "Имя родителя"

    def mailing_view(self, request):
        from django.http import JsonResponse
        from django.urls import reverse
        from .services import CoreService
        from .models import TestResult
        from django.core.files.storage import FileSystemStorage
        import os
        import json
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            action = request.POST.get('action')
            if action == 'calc_recipients':
                age_filter = request.POST.get('age')
                recipients_qs = User.objects.filter(is_staff=False, is_active=True).exclude(
                    telegram_chat_id__isnull=True).exclude(telegram_chat_id='')
                if age_filter and age_filter.strip():
                    try:
                        age = int(age_filter)
                        user_ids_with_age = TestResult.objects.filter(child_age=age).values_list('user_id', flat=True)
                        recipients_qs = recipients_qs.filter(pk__in=user_ids_with_age)
                    except ValueError:
                        pass
                else:
                    recipients_qs = recipients_qs.filter(consent_given=True)
                count = recipients_qs.distinct().count()
                return JsonResponse({'status': 'ok', 'count': count})
            if action == 'upload_temp':
                files = request.FILES.getlist('media_files')
                saved_paths = []
                for f in files:
                    if f.size > 50 * 1024 * 1024:
                        return JsonResponse({'status': 'error', 'message': f'Файл {f.name} больше 50 МБ!'})
                if files:
                    fs = FileSystemStorage(location='media/temp_mailing')
                    for f in files:
                        filename = fs.save(f.name, f)
                        saved_paths.append(fs.path(filename))
                return JsonResponse({'status': 'ok', 'paths': saved_paths})
            if action == 'get_file_ids':
                media_paths_json = request.POST.get('media_paths', '[]')
                message = request.POST.get('message', 'Предпросмотр рассылки')
                paths = json.loads(media_paths_json)
                if not request.user.telegram_chat_id:
                    for p in paths:
                        if os.path.exists(p) and 'temp_mailing' in p:
                            os.remove(p)
                    return JsonResponse({'status': 'error', 'message': 'У вас (Администратора) не подключен Telegram.'})
                try:
                    success, response = CoreService.send_telegram_message(
                        chat_id=request.user.telegram_chat_id,
                        message=f"🛠 <b>ПРОВЕРКА РАССЫЛКИ</b>\n\n{message}",
                        media_paths=paths
                    )
                    for p in paths:
                        if os.path.exists(p) and 'temp_mailing' in p:
                            os.remove(p)
                    if not success:
                        return JsonResponse({'status': 'error', 'message': f'Ошибка Telegram: {response}'})
                    extracted_ids = []
                    if isinstance(response, list):
                        for msg in response:
                            if msg.content_type == 'photo':
                                extracted_ids.append({'type': 'photo', 'id': msg.photo[-1].file_id})
                            elif msg.content_type == 'video':
                                extracted_ids.append({'type': 'video', 'id': msg.video.file_id})
                            elif msg.content_type == 'document':
                                extracted_ids.append({'type': 'document', 'id': msg.document.file_id})
                    else:
                        if response.content_type == 'photo':
                            extracted_ids.append({'type': 'photo', 'id': response.photo[-1].file_id})
                        elif response.content_type == 'video':
                            extracted_ids.append({'type': 'video', 'id': response.video.file_id})
                        elif response.content_type == 'document':
                            extracted_ids.append({'type': 'document', 'id': response.document.file_id})
                    return JsonResponse({'status': 'ok', 'media_ids': extracted_ids})
                except Exception as e:
                    for p in paths:
                        if os.path.exists(p) and 'temp_mailing' in p:
                            os.remove(p)
                    return JsonResponse({'status': 'error', 'message': str(e)})
            if action == 'get_recipients':
                user_ids_raw = request.POST.get('user_ids_hidden')
                age_filter = request.POST.get('age')
                recipients_qs = User.objects.filter(is_staff=False, is_active=True).exclude(
                    telegram_chat_id__isnull=True).exclude(telegram_chat_id='')
                if user_ids_raw:
                    u_ids = [int(x) for x in user_ids_raw.split(',') if x]
                    recipients_qs = recipients_qs.filter(pk__in=u_ids)
                elif age_filter and age_filter.strip():
                    try:
                        age = int(age_filter)
                        user_ids_with_age = TestResult.objects.filter(child_age=age).values_list('user_id', flat=True)
                        recipients_qs = recipients_qs.filter(pk__in=user_ids_with_age)
                    except ValueError:
                        pass
                else:
                    recipients_qs = recipients_qs.filter(consent_given=True)
                id_list = list(recipients_qs.values_list('pk', flat=True))
                if not id_list:
                    return JsonResponse({'status': 'error', 'message': 'Нет пользователей для отправки.'})
                return JsonResponse({'status': 'ok', 'count': len(id_list), 'ids': id_list})
            if action == 'send_single':
                user_id = request.POST.get('user_id')
                message = request.POST.get('message')
                media_ids_json = request.POST.get('media_ids', '[]')
                preview_url = request.POST.get('preview_url')
                buttons_json = request.POST.get('buttons', '[]')
                media_ids = json.loads(media_ids_json)
                buttons = json.loads(buttons_json)
                try:
                    user = User.objects.get(pk=user_id)
                    success = False
                    status_code = "ERROR"
                    if user.telegram_chat_id:
                        success, status = CoreService.send_telegram_message(
                            chat_id=user.telegram_chat_id,
                            message=message,
                            media_ids=media_ids,
                            preview_url=preview_url,
                            buttons=buttons
                        )
                        status_code = "OK" if success else status
                    else:
                        status_code = "NO_ID"
                    return JsonResponse({'status': 'ok', 'sent': success, 'code': status_code})
                except Exception as e:
                    return JsonResponse({'status': 'error', 'code': str(e)})
        context = {
            **self.admin_site.each_context(request),
            'opts': self.model._meta,
            'is_action': False,
            'title': 'Центр рассылок',
            'ajax_url': reverse('admin:user_mailing')
        }
        return render(request, 'admin/mailing.html', context)

    def telegram_message_button(self, obj):
        from django.urls import reverse
        if not obj.pk:
            return "Доступно после сохранения"
        chat_id = getattr(obj, 'telegram_chat_id', None)
        if not chat_id:
            return format_html('<span style="color: #9ca3af; font-weight: bold;">❌ Telegram не подключен</span>')
        url = reverse('admin:user_personal_mailing', args=[obj.pk])
        return format_html(
            '<a href="{}" class="button" style="background-color: #2563eb; color: white; padding: 10px 20px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">'
            '✉️ НАПИСАТЬ СООБЩЕНИЕ (Telegram)</a>',
            url
        )

    telegram_message_button.short_description = "Отправка сообщения"
    telegram_message_button.allow_tags = True

    def personal_mailing_view(self, request, object_id):
        from django.shortcuts import get_object_or_404, render
        from django.urls import reverse
        from .models import TestResult
        user = get_object_or_404(self.model, pk=object_id)
        last_result = TestResult.objects.filter(user=user).order_by('-created_at').first()
        parent_name = user.first_name
        if last_result and last_result.parent_name:
            parent_name = last_result.parent_name
        if not parent_name:
            parent_name = "Не указано"
        recipient_info = {
            'parent_name': parent_name,
            'phone': user.phone,
            'child_name': last_result.child_name if last_result else "Нет данных",
            'child_age': last_result.child_age if last_result else None
        }
        context = {
            **self.admin_site.each_context(request),
            'users': [user],
            'selected_ids': str(user.pk),
            'opts': self.model._meta,
            'is_action': True,
            'recipient_info': recipient_info,
            'title': f'Сообщение пользователю: {user.phone}',
            'ajax_url': reverse('admin:user_mailing')
        }
        return render(request, 'admin/mailing.html', context)

original_index = admin.site.index


def patch_get_app_list():
    from django.contrib import admin
    from django.urls import reverse

    original_get_app_list = admin.site.get_app_list

    def new_get_app_list(request, app_label=None):
        app_list = original_get_app_list(request, app_label)

        testing_order = [
            'AppConfig',
            'TestSuite',
            'TestCategory',
            'Question',
            'FinalQuestion',
            'ReliabilityRule',
            'ProfileMatrix',
            'ProfileRecommendation',
            'KnowledgeBaseDocument',
            'TestResult',
            'Payment',
            'PromoCode',
            'ReportDelivery',
            'AIErrorLog',
            'LandingPage',
            'UtmCampaign',
            'UserUtm',
            'Flower'
        ]

        for app in app_list:
            if app['app_label'] == 'users':
                has_mailing = any(m['object_name'] == 'Mailing' for m in app['models'])
                if not has_mailing:
                    app['models'].append({
                        'name': 'Центр рассылок',
                        'object_name': 'Mailing',
                        'admin_url': reverse('admin:user_mailing'),
                        'add_url': None,
                        'view_only': True
                    })

            if app['app_label'] == 'testing':
                has_flower = any(m['object_name'] == 'Flower' for m in app['models'])
                if not has_flower:
                    app['models'].append({
                        'name': 'Мониторинг (Flower)',
                        'object_name': 'Flower',
                        'admin_url': 'http://84.54.28.169:5555/',
                        'add_url': None,
                        'view_only': True
                    })

                app['models'].sort(
                    key=lambda x: testing_order.index(x['object_name']) if x['object_name'] in testing_order else 99)

        return app_list

    admin.site.get_app_list = new_get_app_list

patch_get_app_list()


@admin.register(AIErrorLog)
class AIErrorLogAdmin(admin.ModelAdmin):
    list_display = ('error_type', 'model_used', 'temp_used', 'created_at', 'result')
    list_filter = ('error_type', 'model_used', 'created_at')
    readonly_fields = ('error_type', 'message', 'model_used', 'temp_used', 'raw_response', 'created_at', 'result')

    def has_add_permission(self, request):
        return False


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = ('id', 'get_child_name', 'get_result_link', 'get_user_link', 'amount', 'status_badge', 'created_at', 'confirmed_at')
    list_filter = ('status', 'created_at')
    search_fields = ('operation_id', 'result__child_name', 'result__user__phone')
    readonly_fields = (
        'result', 'operation_id', 'amount', 'status', 'payment_url',
        'created_at', 'confirmed_at'
    )
    list_per_page = 30
    ordering = ['-created_at']

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_child_name(self, obj):
        return obj.result.child_name if obj.result else '—'
    get_child_name.short_description = 'Ребёнок'

    def get_result_link(self, obj):
        if not obj.result:
            return '—'
        from django.urls import reverse
        url = reverse('admin:testing_testresult_change', args=[obj.result.pk])
        return format_html('<a href="{}">Результат #{}</a>', url, obj.result.pk)
    get_result_link.short_description = 'Результат'

    def get_user_link(self, obj):
        if not obj.result or not obj.result.user:
            return '—'
        from django.urls import reverse
        user = obj.result.user
        url = reverse('admin:users_user_change', args=[user.pk])
        return format_html('<a href="{}">{}</a>', url, user.phone or user.pk)
    get_user_link.short_description = 'Пользователь'

    def status_badge(self, obj):
        colors = {
            'pending': '#f59e0b',
            'approved': '#16a34a',
            'failed': '#dc2626',
            'expired': '#6b7280',
        }
        color = colors.get(obj.status, '#6b7280')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; '
            'border-radius:12px; font-size:11px; font-weight:bold;">{}</span>',
            color, obj.get_status_display()
        )
    status_badge.short_description = 'Статус'


@admin.register(PromoCode)
class PromoCodeAdmin(admin.ModelAdmin):
    list_display = ('code', 'discount_type_display', 'uses_display', 'expires_at', 'is_active', 'is_valid_display', 'created_at')
    list_editable = ('is_active',)
    list_filter = ('discount_type', 'is_active')
    search_fields = ('code',)
    readonly_fields = ('used_count', 'created_at', 'display_usages')
    ordering = ['-created_at']

    fieldsets = (
        (None, {
            'fields': ('code', 'discount_type', 'fixed_price', 'max_uses', 'used_count', 'expires_at', 'is_active', 'created_at')
        }),
        ('Кто использовал', {
            'fields': ('display_usages',)
        }),
    )

    def discount_type_display(self, obj):
        if obj.discount_type == 'free':
            return format_html('<span style="color:#16a34a;font-weight:bold;">Бесплатно</span>')
        return format_html('<span style="color:#2563eb;font-weight:bold;">{} ₽</span>', obj.fixed_price)
    discount_type_display.short_description = 'Скидка'

    def uses_display(self, obj):
        color = '#16a34a' if obj.used_count < obj.max_uses else '#dc2626'
        return format_html('<span style="color:{};font-weight:bold;">{}/{}</span>', color, obj.used_count, obj.max_uses)
    uses_display.short_description = 'Использовано'

    def is_valid_display(self, obj):
        return obj.is_valid()
    is_valid_display.short_description = 'Активен'
    is_valid_display.boolean = True

    def display_usages(self, obj):
        from django.urls import reverse
        from django.utils import timezone

        if not obj.pk:
            return 'Доступно после сохранения'

        results = obj.test_results.select_related('user', 'suite').prefetch_related('payments').order_by('-created_at')

        if not results.exists():
            return mark_safe('<span style="color:#9ca3af;font-weight:bold;">Промокод ещё никто не использовал</span>')

        rows = ''
        for r in results:
            result_url = reverse('admin:testing_testresult_change', args=[r.pk])
            user_url = reverse('admin:users_user_change', args=[r.user.pk])

            dt = timezone.localtime(r.created_at).strftime('%d.%m.%Y %H:%M')
            paid_badge = (
                '<span style="background:#16a34a;color:white;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold;">Оплачен</span>'
                if r.is_paid else
                '<span style="background:#f59e0b;color:white;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:bold;">Не оплачен</span>'
            )

            payment = r.payments.filter(status='approved').first() or r.payments.order_by('-created_at').first()
            if payment:
                payment_url = reverse('admin:testing_payment_change', args=[payment.pk])
                payment_link = f'<a href="{payment_url}" style="color:#2563eb;font-weight:bold;">#{payment.pk} ({payment.amount} ₽)</a>'
            else:
                payment_link = '—'

            rows += f"""
            <tr style="border-bottom:1px solid #f3f4f6;">
                <td style="padding:8px 10px;font-size:13px;white-space:nowrap;">{dt}</td>
                <td style="padding:8px 10px;font-size:13px;">
                    <a href="{result_url}" style="color:#2563eb;font-weight:bold;">#{r.pk} {r.child_name}</a>
                    <div style="font-size:11px;color:#6b7280;">{r.child_age} л. · {r.suite.name}</div>
                </td>
                <td style="padding:8px 10px;font-size:13px;">
                    <a href="{user_url}" style="color:#2563eb;font-weight:bold;">{r.user.phone}</a>
                    {f'<div style="font-size:11px;color:#6b7280;">{r.user.first_name}</div>' if r.user.first_name else ''}
                </td>
                <td style="padding:8px 10px;">{paid_badge}</td>
                <td style="padding:8px 10px;font-size:13px;">{payment_link}</td>
            </tr>"""

        html = f"""
        <div style="overflow-x:auto;">
        <table style="width:100%;border-collapse:collapse;font-family:sans-serif;">
            <thead>
                <tr style="background:#f9fafb;border-bottom:2px solid #e5e7eb;">
                    <th style="padding:8px 10px;text-align:left;font-size:11px;font-weight:900;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;">Дата</th>
                    <th style="padding:8px 10px;text-align:left;font-size:11px;font-weight:900;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;">Результат</th>
                    <th style="padding:8px 10px;text-align:left;font-size:11px;font-weight:900;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;">Пользователь</th>
                    <th style="padding:8px 10px;text-align:left;font-size:11px;font-weight:900;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;">Оплата</th>
                    <th style="padding:8px 10px;text-align:left;font-size:11px;font-weight:900;color:#6b7280;text-transform:uppercase;letter-spacing:.05em;">Платёж</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
        </div>"""
        return mark_safe(html)

    display_usages.short_description = 'Использования промокода'


@admin.register(ReportDelivery)
class ReportDeliveryAdmin(admin.ModelAdmin):
    list_display = ('id', 'get_child_name', 'get_result_link', 'get_user_link', 'report_type_display', 'method_display', 'delivered_at', 'get_delivered_by')
    list_filter = ('report_type', 'method')
    date_hierarchy = 'delivered_at'
    readonly_fields = ('result', 'report_type', 'method', 'delivered_at', 'delivered_by')
    search_fields = ('result__child_name',)

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def get_child_name(self, obj):
        return obj.result.child_name if obj.result else '—'
    get_child_name.short_description = 'Ребёнок'

    def get_result_link(self, obj):
        if not obj.result:
            return '—'
        from django.urls import reverse
        url = reverse('admin:testing_testresult_change', args=[obj.result.pk])
        return format_html('<a href="{}">Результат #{}</a>', url, obj.result.pk)
    get_result_link.short_description = 'Результат'

    def get_user_link(self, obj):
        if not obj.result or not obj.result.user:
            return '—'
        from django.urls import reverse
        user = obj.result.user
        url = reverse('admin:users_user_change', args=[user.pk])
        return format_html('<a href="{}">{}</a>', url, user.phone or user.pk)
    get_user_link.short_description = 'Пользователь'

    def report_type_display(self, obj):
        colors = {'short': '#f59e0b', 'full': '#16a34a'}
        color = colors.get(obj.report_type, '#6b7280')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; '
            'border-radius:12px; font-size:11px; font-weight:bold;">{}</span>',
            color, obj.get_report_type_display()
        )
    report_type_display.short_description = 'Тип'

    def method_display(self, obj):
        return obj.get_method_display()
    method_display.short_description = 'Способ'

    def get_delivered_by(self, obj):
        if obj.delivered_by:
            return obj.delivered_by.first_name or obj.delivered_by.phone
        return '—'
    get_delivered_by.short_description = 'Кто отправил'


@admin.register(LandingPage)
class LandingPageAdmin(admin.ModelAdmin):
    def has_add_permission(self, request):
        return not LandingPage.objects.exists()

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        if LandingPage.objects.exists():
            obj = LandingPage.objects.first()
            return redirect('admin:testing_landingpage_change', obj.pk)
        return super().changelist_view(request, extra_context)


@admin.register(UtmCampaign)
class UtmCampaignAdmin(admin.ModelAdmin):
    list_display = ('name', 'utm_source', 'utm_medium', 'utm_campaign', 'registration_count', 'get_ready_link')
    readonly_fields = ('registration_count', 'get_ready_link')

    def registration_count(self, obj):
        from .models import UserUtm
        count = UserUtm.objects.filter(
            utm_source=obj.utm_source,
            utm_medium=obj.utm_medium,
            utm_campaign=obj.utm_campaign
        ).count()
        return format_html('<b style="color: #2563eb; font-size: 14px;">{}</b>', count)

    def get_ready_link(self, obj):
        if not obj.pk:
            return "Появится после сохранения"

        base_url = "https://humanprofi.ru/test/"
        params = f"?utm_source={obj.utm_source}"
        if obj.utm_medium:
            params += f"&utm_medium={obj.utm_medium}"
        if obj.utm_campaign:
            params += f"&utm_campaign={obj.utm_campaign}"

        full_url = f"{base_url}{params}"

        return format_html(
            '<div style="display:flex; align-items:center; gap:10px;">'
            '<input type="text" value="{}" id="utm_link_{}" readonly '
            'style="width:400px; padding:5px; border-radius:4px; border:1px solid #ccc; background:#f9f9f9;">'
            '<button type="button" onclick="navigator.clipboard.writeText(\'{}\'); alert(\'Ссылка скопирована!\')" '
            'style="background:#2563eb; color:white; border:none; padding:6px 12px; border-radius:4px; cursor:pointer; font-weight:bold;">'
            'Копировать</button>'
            '</div>',
            full_url, obj.pk, full_url
        )

    get_ready_link.short_description = "Готовая ссылка для рекламы"
    registration_count.short_description = "Кол-во регистраций"


@admin.register(KnowledgeBaseDocument)
class KnowledgeBaseDocumentAdmin(admin.ModelAdmin):
    list_display = ('title', 'location', 'is_processed', 'created_at')
    list_filter = ('is_processed', 'location')
    readonly_fields = ('is_processed',)

    def save_model(self, request, obj, form, change):
        from django.db import transaction
        from .tasks import process_knowledge_base_document_task
        super().save_model(request, obj, form, change)
        transaction.on_commit(lambda: process_knowledge_base_document_task.delay(obj.pk))


@admin.register(ProfileMatrix)
class ProfileMatrixAdmin(admin.ModelAdmin):
    list_display = ('profile_name', 'suite', 'index_1', 'index_2')
    list_filter = ('suite',)

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name in ["index_1", "index_2"]:
            obj_id = request.resolver_match.kwargs.get('object_id')
            if obj_id:
                matrix_obj = self.get_object(request, obj_id)
                if matrix_obj:
                    kwargs["queryset"] = TestCategory.objects.filter(suite=matrix_obj.suite)
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

@admin.register(ProfileRecommendation)
class ProfileRecommendationAdmin(admin.ModelAdmin):
    list_display = ('profile_name', 'suite')
    list_filter = ('suite',)
    search_fields = ('profile_name', 'description')


@admin.register(Location)
class LocationAdmin(admin.ModelAdmin):
    list_display = ('name', 'order')
    list_editable = ('order',)

def patch_admin_index():
    from django.contrib import admin

    if not hasattr(admin.site, 'original_index_view'):
        admin.site.original_index_view = admin.site.index

    def new_index(request, extra_context=None):
        stats_data = get_test_statistics()

        context = extra_context or {}
        context.update(stats_data)

        return admin.site.original_index_view(request, extra_context=context)

    admin.site.index = new_index


patch_admin_index()